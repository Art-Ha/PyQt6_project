import sys
import os
import hashlib
import sqlite3
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QDialog, QVBoxLayout, QLabel,
    QLineEdit, QPushButton,
    QHBoxLayout, QFileDialog, QInputDialog, QComboBox
)
from PyQt6 import uic
from PyQt6.QtCore import QDate
from openpyxl import Workbook, load_workbook

DB_FILE = "diary.db"


class Database:
    def __init__(self, db_file):
        self.conn = sqlite3.connect(db_file)
        self.conn.execute("PRAGMA foreign_keys = 1;")
        self.create_tables()

    def create_tables(self):
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL
        )
        """)

        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS theme (
            username TEXT PRIMARY KEY,
            dark INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(username) REFERENCES users(username) ON DELETE CASCADE
        )
        """)

        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            category_name TEXT NOT NULL,
            FOREIGN KEY(username) REFERENCES users(username) ON DELETE CASCADE
        )
        """)

        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            date TEXT NOT NULL,
            text TEXT NOT NULL,
            done INTEGER NOT NULL DEFAULT 0,
            category TEXT NOT NULL,
            priority TEXT NOT NULL,
            FOREIGN KEY(username) REFERENCES users(username) ON DELETE CASCADE
        )
        """)
        self.conn.commit()

    def add_user(self, username, password_hash):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO users (username, password_hash) VALUES (?,?)",
                    (username, password_hash))
        self.conn.commit()

    def get_user(self, username):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT username, password_hash FROM users WHERE username=?",
            (username,))
        return cur.fetchone()

    def update_user_password(self, username, new_hash):
        cur = self.conn.cursor()
        cur.execute("UPDATE users SET password_hash=? WHERE username=?",
                    (new_hash, username))
        self.conn.commit()

    def delete_user(self, username):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM users WHERE username=?", (username,))
        self.conn.commit()

    def get_theme(self, username):
        cur = self.conn.cursor()
        cur.execute("SELECT dark FROM theme WHERE username=?", (username,))
        row = cur.fetchone()
        if row is None:
            return False
        return bool(row[0])

    def set_theme(self, username, dark):
        cur = self.conn.cursor()
        cur.execute("SELECT username FROM theme WHERE username=?", (username,))
        if cur.fetchone() is None:
            cur.execute("INSERT INTO theme (username, dark) VALUES (?,?)",
                        (username, 1 if dark else 0))
        else:
            cur.execute("UPDATE theme SET dark=? WHERE username=?",
                        (1 if dark else 0, username))
        self.conn.commit()

    def get_categories(self, username):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT category_name FROM categories WHERE username=? ORDER BY category_name",
            (username,))
        cats = [row[0] for row in cur.fetchall()]
        if "Все категории" not in cats:
            cats.insert(0, "Все категории")
        return cats

    def add_category(self, username, category_name):
        if category_name == "Все категории":
            return
        cur = self.conn.cursor()
        cur.execute(
            "SELECT category_name FROM categories WHERE username=? AND category_name=?",
            (username, category_name))
        if cur.fetchone() is None:
            cur.execute(
                "INSERT INTO categories (username, category_name) VALUES (?,?)",
                (username, category_name))
            self.conn.commit()

    def delete_category(self, username, category_name):
        if category_name == "Все категории":
            return
        cur = self.conn.cursor()
        cur.execute("DELETE FROM tasks WHERE username=? AND category=?",
                    (username, category_name))
        cur.execute(
            "DELETE FROM categories WHERE username=? AND category_name=?",
            (username, category_name))
        self.conn.commit()

    def add_task(self, username, date_str, text, category, priority):
        if not category:
            category = "Все категории"
        cur = self.conn.cursor()
        cur.execute(
            "INSERT INTO tasks (username, date, text, done, category, priority) VALUES (?,?,?,?,?,?)",
            (username, date_str, text, 0, category, priority)
        )
        self.conn.commit()

    def get_tasks_for_date(self, username, date_str):
        cur = self.conn.cursor()
        cur.execute(
            "SELECT text, done, category, priority FROM tasks WHERE username=? AND date=?",
            (username, date_str))
        return cur.fetchall()

    def delete_task(self, username, date_str, text, done, category, priority):
        cur = self.conn.cursor()
        d_val = 1 if done else 0
        cur.execute(
            "DELETE FROM tasks WHERE username=? AND date=? AND text=? AND done=? AND category=? AND priority=?",
            (username, date_str, text, d_val, category, priority))
        self.conn.commit()

    def update_task_done(self, username, date_str, text, category, priority,
                         done_state):
        d_val = 1 if done_state else 0
        cur = self.conn.cursor()
        cur.execute("""UPDATE tasks 
                       SET done=? 
                       WHERE username=? AND date=? AND text=? AND category=? AND priority=?""",
                    (d_val, username, date_str, text, category, priority))
        self.conn.commit()

    def delete_all_done_tasks(self, username, date_str):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM tasks WHERE username=? AND date=? AND done=1",
                    (username, date_str))
        self.conn.commit()

    def mark_all_tasks_done(self, username, date_str):
        cur = self.conn.cursor()
        cur.execute("UPDATE tasks SET done=1 WHERE username=? AND date=?",
                    (username, date_str))
        self.conn.commit()

    def get_stats(self, username):
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*), SUM(done) FROM tasks WHERE username=?",
                    (username,))
        row = cur.fetchone()
        total_tasks = row[0]
        done_tasks = row[1] if row[1] is not None else 0
        return total_tasks, done_tasks

    def get_monthly_tasks(self, username, month):
        month_str = f"{month:02d}"
        cur = self.conn.cursor()
        cur.execute(
            "SELECT date, text, done, category, priority FROM tasks WHERE username=? AND strftime('%m', date)=?",
            (username, month_str))
        return cur.fetchall()

    def export_categories(self, username, filename):
        cats = self.get_categories(username)
        with open(filename, "w", encoding="utf-8") as f:
            for c in cats:
                if c != "Все категории":
                    f.write(c + "\n")

    def import_categories(self, username, filename):
        if not os.path.exists(filename):
            return
        with open(filename, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()
        for c in lines:
            cat = c.strip()
            if cat:
                self.add_category(username, cat)


class LoginDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Авторизация")
        self.username_edit = QLineEdit()
        self.username_edit.setPlaceholderText("Логин")
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_edit.setPlaceholderText("Пароль")
        self.login_button = QPushButton("Войти")
        self.register_button = QPushButton("Зарегистрироваться")
        self.cancel_button = QPushButton("Отмена")
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Введите логин и пароль:"))
        layout.addWidget(self.username_edit)
        layout.addWidget(self.password_edit)
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.login_button)
        btn_layout.addWidget(self.register_button)
        btn_layout.addWidget(self.cancel_button)
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        self.login_button.clicked.connect(self.attempt_login)
        self.register_button.clicked.connect(self.attempt_register)
        self.cancel_button.clicked.connect(self.reject)
        self.logged_in_username = None

    def attempt_login(self):
        username = self.username_edit.text().strip()
        password = self.password_edit.text().strip()
        if not username or not password:
            QMessageBox.warning(self, "Ошибка", "Введите логин и пароль!")
            return
        user = self.db.get_user(username)
        if user is None:
            QMessageBox.warning(self, "Ошибка", "Пользователь не найден!")
            return
        stored_hash = user[1]
        hashed = hashlib.sha256(password.encode('utf-8')).hexdigest()
        if stored_hash == hashed:
            self.logged_in_username = username
            self.accept()
        else:
            QMessageBox.warning(self, "Ошибка", "Неверный пароль!")

    def attempt_register(self):
        username = self.username_edit.text().strip()
        password = self.password_edit.text().strip()
        if not username or not password:
            QMessageBox.warning(self, "Ошибка", "Введите логин и пароль!")
            return
        user = self.db.get_user(username)
        if user is not None:
            QMessageBox.warning(self, "Ошибка", "Пользователь уже существует!")
            return
        hashed = hashlib.sha256(password.encode('utf-8')).hexdigest()
        self.db.add_user(username, hashed)
        QMessageBox.information(self, "Успех",
                                "Пользователь успешно зарегистрирован!")
        self.logged_in_username = username
        self.accept()

    def get_username(self):
        return self.logged_in_username


class ChangePasswordDialog(QDialog):
    def __init__(self, db, username, parent=None):
        super().__init__(parent)
        self.db = db
        self.username = username
        self.setWindowTitle("Изменить пароль")
        self.old_password_edit = QLineEdit()
        self.old_password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.old_password_edit.setPlaceholderText("Старый пароль")
        self.new_password_edit = QLineEdit()
        self.new_password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.new_password_edit.setPlaceholderText("Новый пароль")
        self.confirm_edit = QLineEdit()
        self.confirm_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.confirm_edit.setPlaceholderText("Подтверждение нового пароля")
        self.ok_button = QPushButton("Изменить")
        self.cancel_button = QPushButton("Отмена")
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Старый пароль:"))
        layout.addWidget(self.old_password_edit)
        layout.addWidget(QLabel("Новый пароль:"))
        layout.addWidget(self.new_password_edit)
        layout.addWidget(QLabel("Подтвердите новый пароль:"))
        layout.addWidget(self.confirm_edit)
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.ok_button)
        btn_layout.addWidget(self.cancel_button)
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        self.ok_button.clicked.connect(self.change_password)
        self.cancel_button.clicked.connect(self.reject)

    def change_password(self):
        old_p = self.old_password_edit.text().strip()
        new_p = self.new_password_edit.text().strip()
        conf_p = self.confirm_edit.text().strip()
        if not old_p or not new_p or not conf_p:
            QMessageBox.warning(self, "Ошибка",
                                "Все поля должны быть заполнены!")
            return
        user = self.db.get_user(self.username)
        if user is None:
            QMessageBox.warning(self, "Ошибка", "Пользователь не найден!")
            return
        old_hash = hashlib.sha256(old_p.encode('utf-8')).hexdigest()
        if user[1] != old_hash:
            QMessageBox.warning(self, "Ошибка", "Старый пароль неверен!")
            return
        if new_p != conf_p:
            QMessageBox.warning(self, "Ошибка", "Пароли не совпадают!")
            return
        new_hash = hashlib.sha256(new_p.encode('utf-8')).hexdigest()
        self.db.update_user_password(self.username, new_hash)
        QMessageBox.information(self, "Успех", "Пароль успешно изменен!")
        self.accept()


class MainWindow(QMainWindow):
    def __init__(self, db, username):
        super().__init__()
        uic.loadUi("mainwindow.ui", self)
        self.db = db
        self.current_user = username

        self.priorityFilterComboBox = QComboBox()
        self.priorityFilterComboBox.addItem("Все приоритеты")
        self.priorityFilterComboBox.addItem("Низкий")
        self.priorityFilterComboBox.addItem("Средний")
        self.priorityFilterComboBox.addItem("Высокий")
        self.gridLayout.addWidget(self.priorityFilterComboBox, 0, 5)
        self.priorityFilterComboBox.currentIndexChanged.connect(
            self.update_task_list)

        self.changePriorityButton = QPushButton("Изменить приоритет")
        self.gridLayout.addWidget(self.changePriorityButton, 8, 1)
        self.changePriorityButton.clicked.connect(
            self.change_selected_task_priority)

        self.changeCategoryButton = QPushButton("Изменить категорию")
        self.gridLayout.addWidget(self.changeCategoryButton, 8, 2)
        self.changeCategoryButton.clicked.connect(
            self.change_selected_task_category)

        self.add_category_combobox()
        self.add_category_button = QPushButton("Добавить категорию")
        self.delete_category_button = QPushButton("Удалить категорию")
        self.gridLayout.addWidget(self.add_category_button, 7, 1)
        self.gridLayout.addWidget(self.delete_category_button, 7, 2)
        self.add_category_button.clicked.connect(self.add_category)
        self.delete_category_button.clicked.connect(self.delete_category)

        self.addTaskButton.clicked.connect(self.add_task)
        self.deleteTaskButton.clicked.connect(self.delete_task)
        self.markDoneButton.clicked.connect(self.mark_task_done)
        self.unmarkButton.clicked.connect(self.unmark_task)
        self.statsButton.clicked.connect(self.show_stats)
        self.calendarWidget.selectionChanged.connect(self.update_task_list)
        self.searchLineEdit.textChanged.connect(self.update_task_list)
        self.deleteAllDoneButton.clicked.connect(self.delete_all_done_tasks)
        self.markAllDoneButton.clicked.connect(self.mark_all_tasks_done)
        self.actionSave_to_Excel.triggered.connect(self.save_to_excel)
        self.actionLoad_from_Excel.triggered.connect(self.load_from_excel)
        self.actionExport_Stats_to_Excel.triggered.connect(
            self.export_stats_to_excel)
        self.actionLogout.triggered.connect(self.logout)
        self.actionAbout.triggered.connect(self.show_about_dialog)
        self.actionDelete_User = self.menuFile.addAction(
            "Удалить пользователя")
        self.actionDelete_User.triggered.connect(self.delete_user)
        self.actionPrint_Month = self.menuFile.addAction(
            "Показать задачи за месяц")
        self.actionPrint_Month.triggered.connect(self.print_monthly_tasks)
        self.actionSave_to_Text = self.menuFile.addAction("Сохранить в текст")
        self.actionSave_to_Text.triggered.connect(self.save_to_text)

        self.sort_by_date = False
        self.actionSort = self.menuFile.addAction("Сортировать задачи по дате")
        self.actionSort.triggered.connect(self.toggle_sort_by_date)

        self.actionChangePassword = self.menuFile.addAction("Изменить пароль")
        self.actionChangePassword.triggered.connect(self.change_password)
        self.actionExportCategories = self.menuFile.addAction(
            "Экспорт категорий")
        self.actionExportCategories.triggered.connect(self.export_categories)
        self.actionImportCategories = self.menuFile.addAction(
            "Импорт категорий")
        self.actionImportCategories.triggered.connect(self.import_categories)
        self.actionToggleTheme = self.menuFile.addAction("Переключить тему")
        self.actionToggleTheme.triggered.connect(self.toggle_theme)
        self.setWindowTitle(f"Ежедневник - Пользователь: {self.current_user}")
        self.update_category_list()
        self.load_theme()
        self.update_task_list()

    def toggle_sort_by_date(self):
        self.sort_by_date = not self.sort_by_date
        self.update_task_list()

    def load_theme(self):
        dark = self.db.get_theme(self.current_user)
        if dark:
            self.set_dark_theme()
        else:
            self.set_light_theme()

    def set_dark_theme(self):
        self.setStyleSheet(
            "QWidget { background-color: #2B2B2B; color: #EEEEEE; }")

    def set_light_theme(self):
        self.setStyleSheet("")

    def toggle_theme(self):
        current = self.db.get_theme(self.current_user)
        new_val = not current
        self.db.set_theme(self.current_user, new_val)
        if new_val:
            self.set_dark_theme()
        else:
            self.set_light_theme()

    def add_category_combobox(self):
        self.categoryComboBox = QComboBox()
        self.categoryComboBox.currentIndexChanged.connect(
            self.update_task_list)
        self.gridLayout.addWidget(self.categoryComboBox, 0, 4)

    def add_category(self):
        category, ok = QInputDialog.getText(self, "Добавить категорию",
                                            "Название категории:")
        if ok and category.strip():
            self.db.add_category(self.current_user, category.strip())
            self.update_category_list()

    def delete_category(self):
        if self.categoryComboBox.currentText() == "Все категории":
            QMessageBox.warning(self, "Ошибка",
                                "Эту категорию нельзя удалить.")
            return
        cat = self.categoryComboBox.currentText()
        self.db.delete_category(self.current_user, cat)
        self.update_category_list()
        self.update_task_list()

    def update_category_list(self):
        self.categoryComboBox.clear()
        categories = self.db.get_categories(self.current_user)
        for c in categories:
            self.categoryComboBox.addItem(c)

    def get_selected_date(self):
        return self.calendarWidget.selectedDate()

    def add_task(self):
        task_text = self.taskLineEdit.text().strip()
        if not task_text:
            QMessageBox.warning(self, "Ошибка", "Введите текст задачи!")
            return
        date = self.get_selected_date()
        cat = self.categoryComboBox.currentText()
        priority = self.priorityComboBox.currentText()
        date_str = date.toString("yyyy-MM-dd")
        self.db.add_task(self.current_user, date_str, task_text, cat, priority)
        self.taskLineEdit.clear()
        self.update_task_list()

    def delete_task(self):
        selected_items = self.tasksListWidget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Ошибка",
                                "Выберите задачу для удаления!")
            return
        date = self.get_selected_date()
        date_str = date.toString("yyyy-MM-dd")
        selected_item = selected_items[0].text()
        done = selected_item.startswith("[✓]")
        display_str = selected_item.replace("[✓] ", "",
                                            1) if done else selected_item

        cat_start = display_str.rfind("(")
        cat_end = display_str.rfind(")")
        priority_start = display_str.rfind("[")
        priority_end = display_str.rfind("]")
        text = display_str[:cat_start].strip()
        cat = display_str[cat_start + 1:cat_end]
        prio = display_str[priority_start + 1:priority_end]

        self.db.delete_task(self.current_user, date_str, text, done, cat, prio)
        self.update_task_list()

    def mark_task_done(self):
        self._set_task_done_state(True)

    def unmark_task(self):
        self._set_task_done_state(False)

    def _set_task_done_state(self, done_state: bool):
        selected_items = self.tasksListWidget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Ошибка", "Выберите задачу!")
            return
        date = self.get_selected_date()
        date_str = date.toString("yyyy-MM-dd")
        selected_item = selected_items[0].text()
        done = selected_item.startswith("[✓]")
        display_str = selected_item.replace("[✓] ", "",
                                            1) if done else selected_item

        cat_start = display_str.rfind("(")
        cat_end = display_str.rfind(")")
        priority_start = display_str.rfind("[")
        priority_end = display_str.rfind("]")
        text = display_str[:cat_start].strip()
        cat = display_str[cat_start + 1:cat_end]
        prio = display_str[priority_start + 1:priority_end]

        self.db.update_task_done(self.current_user, date_str, text, cat, prio,
                                 done_state)
        self.update_task_list()

    def filter_tasks(self, tasks_list):
        search_query = self.searchLineEdit.text().strip().lower()
        selected_cat = self.categoryComboBox.currentText()
        selected_prio = self.priorityFilterComboBox.currentText()

        filtered = []
        for (text, done, cat, prio) in tasks_list:
            if selected_cat != "Все категории" and cat != selected_cat:
                continue
            if selected_prio != "Все приоритеты" and prio != selected_prio:
                continue
            if search_query and search_query not in text.lower():
                continue
            filtered.append((text, done, cat, prio))
        return filtered

    def update_task_list(self):
        self.tasksListWidget.clear()
        date = self.get_selected_date()
        date_str = date.toString("yyyy-MM-dd")
        tasks_list = self.db.get_tasks_for_date(self.current_user, date_str)
        filtered = self.filter_tasks(tasks_list)
        for text, done, cat, prio in filtered:
            display_text = (
                               "[✓] " if done else "") + text + f" ({cat}) [{prio}]"
            self.tasksListWidget.addItem(display_text)

    def show_stats(self):
        total_tasks, done_tasks = self.db.get_stats(self.current_user)
        stats_dialog = QDialog(self)
        stats_dialog.setWindowTitle("Статистика")
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"Всего задач: {total_tasks}"))
        layout.addWidget(QLabel(f"Выполнено задач: {done_tasks}"))
        stats_dialog.setLayout(layout)
        stats_dialog.exec()

    def save_to_excel(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Сохранить в Excel",
                                                  "", "Excel Files (*.xlsx)")
        if not filename:
            return
        conn = self.db.conn
        cur = conn.cursor()
        cur.execute(
            "SELECT date, text, done, category, priority FROM tasks WHERE username=? ORDER BY date",
            (self.current_user,))
        rows = cur.fetchall()
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        ws.append(["Дата", "Задача", "Выполнена", "Категория", "Приоритет"])
        for date_val, text, done, cat, prio in rows:
            ws.append([date_val, text, "Да" if done else "Нет", cat, prio])
        wb.save(filename)
        QMessageBox.information(self, "Успех",
                                f"Данные успешно сохранены в {filename}")

    def load_from_excel(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Загрузить из Excel",
                                                  "", "Excel Files (*.xlsx)")
        if not filename:
            return
        if not os.path.exists(filename):
            QMessageBox.warning(self, "Ошибка", "Файл не найден!")
            return
        try:
            wb = load_workbook(filename)
            ws = wb.active
            conn = self.db.conn
            cur = conn.cursor()
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                date_val, text, done_str, cat, prio = row
                if not date_val or not text:
                    continue
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                d = QDate.fromString(date_str, "yyyy-MM-dd")
                if not d.isValid():
                    continue
                done = (done_str == "Да")
                if not cat:
                    cat = "Все категории"
                self.db.add_category(self.current_user, cat)
                if not prio:
                    prio = "Низкий"
                cur.execute(
                    "INSERT INTO tasks (username, date, text, done, category, priority) VALUES (?,?,?,?,?,?)",
                    (self.current_user, date_str, text, 1 if done else 0, cat,
                     prio))
            conn.commit()
            self.update_category_list()
            self.update_task_list()
            QMessageBox.information(self, "Успех",
                                    f"Данные успешно загружены из {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка",
                                 f"Не удалось загрузить файл: {e}")

    def export_stats_to_excel(self):
        filename, _ = QFileDialog.getSaveFileName(self,
                                                  "Экспорт статистики в Excel",
                                                  "", "Excel Files (*.xlsx)")
        if not filename:
            return
        total_tasks, done_tasks = self.db.get_stats(self.current_user)
        wb = Workbook()
        ws = wb.active
        ws.title = "Stats"
        ws.append(["Всего задач", "Выполнено задач"])
        ws.append([total_tasks, done_tasks])
        wb.save(filename)
        QMessageBox.information(self, "Успех",
                                f"Статистика успешно сохранена в {filename}")

    def logout(self):
        self.close()
        main()

    def show_about_dialog(self):
        QMessageBox.information(self, "О программе",
                                "Ежедневник\nВерсия 3.0\nПоддержка пользователей, категорий, приоритета, импорта/экспорта категорий, смена темы.")

    def delete_user(self):
        reply = QMessageBox.question(self, "Удалить пользователя",
                                     "Вы уверены, что хотите удалить этого пользователя и все его данные?")
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_user(self.current_user)
            QMessageBox.information(self, "Успех",
                                    "Пользователь успешно удален.")
            self.logout()

    def print_monthly_tasks(self):
        month, ok = QInputDialog.getInt(self, "Выбор месяца",
                                        "Введите номер месяца (1-12):",
                                        datetime.now().month, 1,
                                        12)
        if not ok:
            return
        rows = self.db.get_monthly_tasks(self.current_user, month)
        text_result = ""
        for (date_val, text, done, cat, prio) in rows:
            status = "[✓]" if done else "[ ]"
            text_result += f"{date_val}: {status} {text} ({cat}) [{prio}]\n"
        dlg = QDialog(self)
        dlg.setWindowTitle("Задачи за месяц")
        v = QVBoxLayout()
        v.addWidget(QLabel(
            text_result if text_result else "Нет задач за выбранный месяц."))
        dlg.setLayout(v)
        dlg.exec()

    def save_to_text(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Сохранить в текст",
                                                  "", "Text Files (*.txt)")
        if not filename:
            return
        conn = self.db.conn
        cur = conn.cursor()
        cur.execute(
            "SELECT date, text, done, category, priority FROM tasks WHERE username=? ORDER BY date",
            (self.current_user,))
        rows = cur.fetchall()
        lines = []
        for date_val, text, done, cat, prio in rows:
            status = "[✓]" if done else "[ ]"
            lines.append(f"{date_val} | {status} {text} ({cat}) [{prio}]")
        with open(filename, "w", encoding="utf-8") as f:
            for line in lines:
                f.write(line + "\n")
        QMessageBox.information(self, "Успех",
                                f"Данные успешно сохранены в {filename}")

    def change_password(self):
        dlg = ChangePasswordDialog(self.db, self.current_user, self)
        dlg.exec()

    def export_categories(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Экспорт категорий",
                                                  "", "Text Files (*.txt)")
        if not filename:
            return
        self.db.export_categories(self.current_user, filename)
        QMessageBox.information(self, "Успех",
                                f"Категории успешно экспортированы в {filename}")

    def import_categories(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Импорт категорий", "",
                                                  "Text Files (*.txt)")
        if not filename:
            return
        self.db.import_categories(self.current_user, filename)
        self.update_category_list()
        self.update_task_list()
        QMessageBox.information(self, "Успех",
                                f"Категории успешно импортированы из {filename}")

    def delete_all_done_tasks(self):
        date = self.get_selected_date()
        date_str = date.toString("yyyy-MM-dd")
        self.db.delete_all_done_tasks(self.current_user, date_str)
        self.update_task_list()

    def mark_all_tasks_done(self):
        date = self.get_selected_date()
        date_str = date.toString("yyyy-MM-dd")
        self.db.mark_all_tasks_done(self.current_user, date_str)
        self.update_task_list()

    def change_selected_task_priority(self):
        selected_items = self.tasksListWidget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Ошибка", "Выберите задачу!")
            return

        selected_item = selected_items[0].text()
        done = selected_item.startswith("[✓]")
        display_str = selected_item.replace("[✓] ", "",
                                            1) if done else selected_item

        cat_start = display_str.rfind("(")
        cat_end = display_str.rfind(")")
        priority_start = display_str.rfind("[")
        priority_end = display_str.rfind("]")
        text = display_str[:cat_start].strip()
        cat = display_str[cat_start + 1:cat_end]
        old_prio = display_str[priority_start + 1:priority_end]

        new_prio, ok = QInputDialog.getItem(self, "Изменить приоритет",
                                            "Выберите новый приоритет:",
                                            ["Низкий", "Средний", "Высокий"],
                                            0, False)
        if not ok:
            return

        self.update_task_in_db(text, done, cat, old_prio, new_prio=new_prio)
        self.update_task_list()

    def change_selected_task_category(self):
        selected_items = self.tasksListWidget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Ошибка", "Выберите задачу!")
            return

        selected_item = selected_items[0].text()
        done = selected_item.startswith("[✓]")
        display_str = selected_item.replace("[✓] ", "",
                                            1) if done else selected_item

        cat_start = display_str.rfind("(")
        cat_end = display_str.rfind(")")
        priority_start = display_str.rfind("[")
        priority_end = display_str.rfind("]")
        text = display_str[:cat_start].strip()
        old_cat = display_str[cat_start + 1:cat_end]
        prio = display_str[priority_start + 1:priority_end]

        cats = self.db.get_categories(self.current_user)
        new_cat, ok = QInputDialog.getItem(self, "Изменить категорию",
                                           "Выберите новую категорию:", cats,
                                           0, False)
        if not ok or not new_cat:
            return

        self.update_task_in_db(text, done, old_cat, prio, new_cat=new_cat)
        self.update_task_list()

    def update_task_in_db(self, text, done, cat, prio, new_cat=None,
                          new_prio=None):
        d_val = 1 if done else 0
        date_str = self.get_selected_date().toString("yyyy-MM-dd")
        conn = self.db.conn
        cur = conn.cursor()

        updates = []
        params = []
        if new_cat is not None:
            updates.append("category=?")
            params.append(new_cat)
        if new_prio is not None:
            updates.append("priority=?")
            params.append(new_prio)

        if not updates:
            return

        set_clause = ", ".join(updates)
        params.extend([self.current_user, date_str, text, d_val, cat, prio])
        query = f"UPDATE tasks SET {set_clause} WHERE username=? AND date=? AND text=? AND done=? AND category=? AND priority=?"
        cur.execute(query, tuple(params))
        conn.commit()


def main():
    app = QApplication(sys.argv)
    db = Database(DB_FILE)
    login_dialog = LoginDialog(db)
    if login_dialog.exec() == QDialog.DialogCode.Accepted:
        username = login_dialog.get_username()
        window = MainWindow(db, username)
        window.show()
        sys.exit(app.exec())
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
